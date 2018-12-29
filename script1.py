import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

print "Opening Browser"
driver = webdriver.Chrome()
# Give the location of the file
path = "./excel/whitepapers.xlsx"

# workbook object is created
wb_obj = openpyxl.load_workbook(path)

sheet_obj = wb_obj.active
m_row = sheet_obj.max_row
print "Number of rows identified: "+str(m_row);
# Loop will print all values
# of first column

count=1
for i in range(2, m_row + 1):
    print "Loop: "+ str(count)
    count += 1
    cell_obj = sheet_obj.cell(row=i, column=5)
    print("Image Link: "+str(cell_obj.value))

    if cell_obj.value and not sheet_obj.cell(row=i,column=6).value:
        try:
            driver.get(cell_obj.value)
            #driver.minimize_window()
            image=WebDriverWait(driver,30).until(EC.presence_of_element_located((By.CLASS_NAME,"od-ImageTile-image")))
            print "Image URL: "+str(image.get_attribute('src'))
            sheet_obj.cell(row=i, column=6).value = image.get_attribute('src')
            print "Don't close...Saving to excel...file may corrupted"
            wb_obj.save(path)
            print "Saved changes in Excel file:" + str(path)
        except:
            print "Something went wrong...page takes a while to open"
    else:
        continue

print "Closing Browser"
driver.quit()